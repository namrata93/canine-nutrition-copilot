"""Print a structural inventory of the Dog Formulator workbook.

Usage:
    python scripts/inspect_workbook.py data/dog-formulator.xlsx
"""

from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
import re

from openpyxl import load_workbook

SHEET_REFERENCE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9 _-]+))!")


def inspect(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)

    print(f"Workbook: {path}")
    print(f"Sheets: {len(workbook.sheetnames)}")

    for sheet in workbook.worksheets:
        formula_count = 0
        references: Counter[str] = Counter()

        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue

                formula_count += 1
                for quoted, unquoted in SHEET_REFERENCE.findall(value):
                    references[(quoted or unquoted).strip()] += 1

        print(
            f"- {sheet.title}: {sheet.max_row} rows x {sheet.max_column} columns, "
            f"{formula_count} formulas"
        )
        if references:
            print(f"  references: {dict(references)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    if not args.workbook.exists():
        parser.error(f"Workbook does not exist: {args.workbook}")

    inspect(args.workbook)
